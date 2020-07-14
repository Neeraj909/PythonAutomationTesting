import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def getReadData(file, sheetName, rouNum, columnNum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rouNum, column=columnNum).value


def writeDataIntoExcel(file, sheetName, rouNum, columnNum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rouNum, column=columnNum).value = data
    workbook.save(file)
